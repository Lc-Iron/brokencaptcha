import os 
import tempfile 
from selenium import webdriver 
from selenium.webdriver.edge.service import Service 
from selenium.webdriver.edge.webdriver import WebDriver 
from selenium.webdriver.common.by import By 
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.common.action_chains import ActionChains 
import requests 
import whisper 
import subprocess 
import re 
import openpyxl 
from datetime import datetime 
import time 
import random 
import unidecode 
 
# Registra as ações no arquivo de log 
def write_log(message): 
    with open("captcha_log.txt", "a", encoding='utf-8') as log_file: 
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S") 
        log_file.write(f"[{timestamp}] {message}\n") 
 
# Simula uma pequena pausa entre ações 
def human_delay(): 
    time.sleep(random.uniform(0.3, 0.7)) 
 
# Digita o texto como se fosse uma pessoa 
def human_type(element, text): 
    for char in text: 
        element.send_keys(char) 
        time.sleep(random.uniform(0.05, 0.1)) 
 
# Move o mouse de forma mais natural 
def move_to_element_human_like(driver, element): 
    action = ActionChains(driver) 
    action.move_to_element(element) 
    action.pause(random.uniform(0.05, 0.1)) 
    action.perform() 
 
# Espera um elemento aparecer na página 
def wait_for_element(driver, locator, timeout=5): 
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located(locator)) 
 
# Normaliza o número do processo 
def normalize_process_number(numero_processo): 
    return unidecode.unidecode(str(numero_processo)) 
 
# Tenta resolver o captcha para um número de processo 
def process_numero(numero_processo, driver, max_tentativas=5): 
    for tentativa in range(1, max_tentativas + 1): 
        try: 
            human_delay() 
 
            # Preenche o campo do processo 
            campo_processo = wait_for_element(driver, (By.ID, "txtProtocoloPesquisa")) 
            move_to_element_human_like(driver, campo_processo) 
            campo_processo.clear() 
            human_type(campo_processo, numero_processo) 
 
            # Pega o link do áudio do captcha 
            src = wait_for_element(driver, (By.XPATH, "/html/body/div[1]/div/div/form/div/div[2]/div[1]/div[2]/audio/source")) 
            link_audio = src.get_attribute("src") 
 
            # Baixa e converte o áudio 
            mp3_file_path, wav_file_path = download_and_convert_audio(link_audio, driver) 
 
            # Transcreve o áudio do captcha 
            captcha = transcribe_audio(wav_file_path) 
 
            # Preenche o captcha 
            campo_captcha = wait_for_element(driver, (By.ID, "txtInfraCaptcha")) 
            move_to_element_human_like(driver, campo_captcha) 
            campo_captcha.clear() 
            human_type(campo_captcha, captcha) 
 
            # Clica pra pesquisar 
            botao_pesquisar = wait_for_element(driver, (By.ID, "sbmPesquisar")) 
            move_to_element_human_like(driver, botao_pesquisar) 
            botao_pesquisar.click() 
 
            # Vê se deu erro 
            try: 
                WebDriverWait(driver, 3).until( 
                    EC.presence_of_element_located((By.CLASS_NAME, "alert-danger")) 
                ) 
                print(f"Errou o captcha na tentativa {tentativa} pro processo {numero_processo}") 
                write_log(f"Tentativa {tentativa} falhou pro processo {numero_processo}. Captcha errado.") 
                if tentativa < max_tentativas: 
                    human_delay() 
                    driver.refresh()  # Tenta de novo 
                    continue 
                else: 
                    write_log(f"Não conseguiu resolver o captcha pro processo {numero_processo} depois de {max_tentativas} tentativas") 
                    return False 
            except: 
                # Se não achou erro, deve ter dado certo 
                write_log(f"Captcha resolvido com sucesso pro processo {numero_processo}") 
                return True 
 
        except Exception as e: 
            error_message = f"Deu ruim na tentativa {tentativa} pro processo {numero_processo}. Erro: {str(e)}" 
            print(error_message) 
            write_log(error_message) 
            if tentativa < max_tentativas: 
                human_delay() 
                driver.refresh()  # Tenta de novo 
                continue 
            else: 
                write_log(f"Não deu pra resolver o captcha pro processo {numero_processo} depois de {max_tentativas} tentativas") 
                return False 
 
    return False 
 
# Baixa e converte o áudio do captcha 
def download_and_convert_audio(link_audio, driver): 
    session = requests.Session() 
    for cookie in driver.get_cookies(): 
        session.cookies.set(cookie['name'], cookie['value']) 
 
    temp_dir = tempfile.gettempdir() 
    mp3_file_path = os.path.join(temp_dir, "temp_audio.mp3") 
    wav_file_path = os.path.join(temp_dir, "temp_audio.wav") 
 
    response = session.get(link_audio, stream=True) 
    if response.status_code == 200: 
        with open(mp3_file_path, 'wb') as file: 
            for chunk in response.iter_content(chunk_size=8192): 
                file.write(chunk) 
 
        subprocess.run(['ffmpeg', '-i', mp3_file_path, '-y', wav_file_path], 
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL) 
 
    return mp3_file_path, wav_file_path 
 
# Transcreve o áudio do captcha 
def transcribe_audio(wav_file_path): 
    modelo = whisper.load_model("base") 
    resposta = modelo.transcribe(wav_file_path, language="pt") 
    return re.sub(r'[^A-Z0-9]', '', resposta['text'].upper()) 
 
# Configuração inicial 
ffmpeg_path = r'C:\ffmpeg\bin' 
os.environ["PATH"] += os.pathsep + ffmpeg_path 
 
service = Service(r"c:\Users\FX4S\Downloads\Python\edgedriver_win64\msedgedriver.exe") 
options = webdriver.EdgeOptions() 
options.add_argument("start-maximized")  # Abre o navegador maximizado 
driver = WebDriver(service=service, options=options) 
 
try: 
    # Lê os processos do Excel 
    wb = openpyxl.load_workbook(r"C:\Users\FX4S\Downloads\Python\Brokencaptcha\teste.xlsx") 
    sheet = wb.active 
     
    # Passa por cada linha do Excel, começando da A2 
    for row in range(2, sheet.max_row + 1): 
        numero_processo = sheet.cell(row=row, column=1).value 
        if numero_processo: 
            numero_processo = normalize_process_number(numero_processo) 
            print(f"Processando: {numero_processo}") 
             
            # Abre a página de pesquisa 
            driver.get("https://sei.anp.gov.br/sei/modulos/pesquisa/md_pesq_processo_pesquisar.php?acao_externa=protocolo_pesquisar&acao_origem_externa=protocolo_pesquisar&id_orgao_acesso_externo=0") 
             
            # Tenta resolver o captcha 
            if process_numero(numero_processo, driver): 
                sheet.cell(row=row, column=2, value="SIM") 
                print(f"Processo {numero_processo} pesquisado com sucesso.") 
            else: 
                sheet.cell(row=row, column=2, value="NÃO") 
                error_message = f"Não deu pra resolver o captcha pro processo {numero_processo} depois de 5 tentativas" 
                print(error_message) 
                write_log(error_message) 
             
            # Salva as mudanças no Excel depois de cada processo 
            wb.save(r"C:\Users\FX4S\Downloads\Python\Brokencaptcha\teste.xlsx") 
 
except Exception as e: 
    error_message = f"Deu um erro geral: {str(e)}" 
    print(error_message) 
    write_log(error_message) 
 
finally: 
    # Salva as mudanças finais no Excel 
    wb.save(r"C:\Users\FX4S\Downloads\Python\Brokencaptcha\teste.xlsx") 
    input("Aperta Enter pra fechar o navegador...") 
    driver.quit() 

